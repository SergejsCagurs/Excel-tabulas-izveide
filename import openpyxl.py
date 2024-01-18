import openpyxl  # Importējiet openpyxl bibliotēku darbam ar Excel failiem
from openpyxl import Workbook  # Importējiet darbgrāmatas klasi, lai izveidotu jaunus Excel failus
import os  # Importējiet OS moduli, lai strādātu ar failu sistēmu

# Definējiet InventoryManager klasi, lai pārvaldītu krājumus
class InventoryManager:
    def __init__(self, filename):
        self.filename = filename  # Excel faila nosaukums, ar kuru strādāt
        # Pārbaudiet, vai fails pastāv, ja ne, izveidojiet jaunu
        if not os.path.exists(filename):
            self.create_new_file()
        # Ielādējiet datus no faila
        self.load_data()

    def create_new_file(self):
        workbook = Workbook()  # Izveidojiet jaunu Excel darbgrāmatu
        sheet = workbook.active  # Aktīvās lapas iegūšana
        sheet.title = "Inventory"  # Lapas nosaukums "Inventārs"
        headers = ["ID", "Produkta nosaukums", "Daudzums", "Vienības cena"]  # Kolonnu galvenes
        sheet.append(headers)  # Galvenes pievienošana lapai
        workbook.save(self.filename)  # Faila saglabāšana

    def load_data(self):
        self.workbook = openpyxl.load_workbook(self.filename)  # Notiek Excel darbgrāmatas ielāde
        self.sheet = self.workbook.active  # Aktīvās lapas iegūšana

    def save_data(self):
        self.workbook.save(self.filename)  # Saglabā izmaiņas failā

    def add_item(self, item_id, name, quantity, price):
        self.sheet.append([item_id, name, quantity, price])  # Produkta ieraksta pievienošana
        self.save_data()  # Faila saglabāšana

    def update_item(self, item_id, name=None, quantity=None, price=None):
        for row in self.sheet.iter_rows(min_row=2, values_only=False):
            if row[0].value == item_id:  # Meklējiet produktu pēc ID
                if name:
                    row[1].value = name  # Atjauniniet nosaukumu, ja norādīts
                if quantity:
                    row[2].value = quantity  # Atjauniniet daudzumu, ja norādīts
                if price:
                    row[3].value = price  # Atjauniniet cenu, ja tā ir iestatīta
                self.save_data()  # Saglabā izmaiņas
                return
        print("Produkts nav atrasts")  # Ziņojums, ja prece nav atrasta

    def get_item_info(self, item_id):
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == item_id:  # Meklējiet produktu pēc ID
                return row  # Meklējiet produktu pēc ID
        return "Produkts nav atrasts"  # Ziņojums, ja prece nav atrasta

# Funkcija lietotāja mijiedarbībai
def user_interface():
    inventory = InventoryManager("inventory.xlsx")  # Izveidojiet krājumu pārvaldnieku
    while True:
        print("\nKrājumu vadības sistēma")
        print("1. Pievienot produktu")
        print("2. Atjaunināt produktu")
        print("3. Iegūstiet informāciju par produktu")
        print("4. Iziet")
        choice = input("Izvēlieties darbību:")

        if choice == '1':
            item_id = input("Ievadiet produkta ID: ")
            name = input("Ievadiet produkta nosaukumu: ")
            quantity = input("Ievadiet daudzumu:")
            price = input("Ievadiet vienības cenu: ")
            inventory.add_item(item_id, name, quantity, price)
            print("Produkts ir veiksmīgi pievienots!")

        elif choice == '2':
            item_id = input("Ievadiet atjaunināmās preces ID:")
            name = input("Ievadiet jaunu nosaukumu (atstājiet tukšu, lai saglabātu veco):")
            quantity = input("Ievadiet jaunu daudzumu (atstājiet tukšu, lai saglabātu veco):")
            price = input("Ievadiet jaunu cenu (atstājiet tukšu, lai saglabātu veco): ")
            inventory.update_item(item_id, name or None, quantity or None, price or None)
            print("Produkta informācija ir veiksmīgi atjaunināta!")

        elif choice == '3':
            item_id = input("Ievadiet produkta ID, lai iegūtu informāciju:")
            item_info = inventory.get_item_info(item_id)
            if item_info != "Produkts nav atrasts":
                print(f"Produkta ID: {item_info[0]}, Vārds: {item_info[1]}, Daudzums: {item_info[2]},Cena: {item_info[3]}")
            else:
                print(item_info)

        elif choice == '4':
            print("Notiek programmas iziešana...")
            break

        else:
            print("Nepareiza izvēle, mēģiniet vēlreiz.")

# Palaidiet lietotāja interfeisu
if __name__ == "__main__":
    user_interface()